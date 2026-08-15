"""
Reports & Export System.

Generates real PDF and Excel documents from the platform's own data —
no third-party report engine, just ReportLab (PDF) and openpyxl (Excel)
laid out from the same models every other page reads from.
"""

import io
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from pymongo.database import Database
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from models.skin_profile import SkinProfile
from models.user import User
from services import (
    assessment_service,
    lifestyle_service,
    product_service,
    progress_service,
    routine_service,
)

PRIMARY_COLOR = colors.HexColor("#4f46e5")
MUTED_COLOR = colors.HexColor("#6b7280")


def _styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            "ReportTitle", parent=styles["Title"], textColor=PRIMARY_COLOR, fontSize=20, spaceAfter=4
        )
    )
    styles.add(
        ParagraphStyle(
            "ReportSubtitle", parent=styles["Normal"], textColor=MUTED_COLOR, fontSize=10, spaceAfter=16
        )
    )
    styles.add(
        ParagraphStyle(
            "SectionHeading", parent=styles["Heading2"], textColor=PRIMARY_COLOR, fontSize=13, spaceBefore=16, spaceAfter=8
        )
    )
    return styles


def _kv_table(rows: list[tuple[str, str]]) -> Table:
    table = Table(rows, colWidths=[5 * cm, 11 * cm])
    table.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("TEXTCOLOR", (0, 0), (0, -1), MUTED_COLOR),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ]
        )
    )
    return table


def build_skin_health_report_pdf(db: Session, mongo_db: Database | None, user_id: uuid.UUID) -> bytes:
    """
    A single comprehensive PDF: profile, latest assessment breakdown,
    current routine, recent lifestyle logs, adherence, and any product
    recommendations — covering the spec's "skin assessment reports",
    "routine reports", "skin health reports", and "product recommendation
    reports" in one cohesive downloadable document.
    """
    user = db.query(User).filter(User.id == user_id).first()
    skin_profile = db.query(SkinProfile).filter(SkinProfile.user_id == user_id).first()
    latest = assessment_service.get_latest_assessment(db, user_id)
    improvement = assessment_service.compute_improvement(db, user_id)
    routine_steps = routine_service.get_active_routine(db, user_id)
    lifestyle_logs = lifestyle_service.list_lifestyle_logs(db, user_id)[:7]
    adherence = progress_service.compute_adherence(db, mongo_db, user_id) if mongo_db is not None else {}
    recommendations = product_service.list_recommendations_for_client(db, user_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = _styles()
    story = []

    story.append(Paragraph("Skin Health Report", styles["ReportTitle"]))
    story.append(
        Paragraph(
            f"{user.full_name if user else 'Unknown user'} · Generated {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}",
            styles["ReportSubtitle"],
        )
    )

    story.append(Paragraph("Skin Profile", styles["SectionHeading"]))
    story.append(
        _kv_table(
            [
                ("Skin type", skin_profile.skin_type if skin_profile else "Not set"),
                ("Skin concerns", skin_profile.skin_concerns if skin_profile and skin_profile.skin_concerns else "Not set"),
                ("Allergies", skin_profile.allergies if skin_profile and skin_profile.allergies else "None reported"),
                ("Sensitivity level", skin_profile.sensitivity_level if skin_profile and skin_profile.sensitivity_level else "Not set"),
            ]
        )
    )

    if latest:
        story.append(Paragraph("Latest Skin Health Score", styles["SectionHeading"]))
        rows = [
            ("Overall score", f"{round(latest.overall_score)} / 100"),
            ("Skin Condition (35%)", f"{round(latest.skin_condition_score, 1)}"),
            ("Lifestyle (20%)", f"{round(latest.lifestyle_score, 1)}"),
            ("Sleep (15%)", f"{round(latest.sleep_score, 1)}"),
            ("Routine Consistency (20%)", f"{round(latest.consistency_score, 1)}"),
            ("Hydration (10%)", f"{round(latest.hydration_score, 1)}"),
            ("Primary concern", latest.primary_concern or "None flagged"),
            ("Assessed on", latest.created_at.strftime("%d %b %Y")),
        ]
        if improvement:
            rows.append(
                (
                    "Improvement",
                    f"{'+' if improvement['delta_points'] > 0 else ''}{improvement['delta_points']} pts "
                    f"({improvement['trend']}) since {improvement['since'].strftime('%d %b %Y')}",
                )
            )
        story.append(_kv_table(rows))

        if latest.detected_concerns:
            story.append(Spacer(1, 10))
            concern_rows = [["Concern", "Severity"]] + [
                [c.get("name", ""), c.get("severity", "")] for c in latest.detected_concerns
            ]
            concern_table = Table(concern_rows, colWidths=[8 * cm, 8 * cm])
            concern_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                    ]
                )
            )
            story.append(concern_table)
    else:
        story.append(Paragraph("Latest Skin Health Score", styles["SectionHeading"]))
        story.append(Paragraph("No assessment completed yet.", styles["Normal"]))

    story.append(Paragraph("Current Routine", styles["SectionHeading"]))
    if routine_steps:
        routine_rows = [["Time", "Step", "Category"]] + [
            [s.time_of_day, str(s.step_number), s.step_category] for s in routine_steps
        ]
        routine_table = Table(routine_rows, colWidths=[4 * cm, 3 * cm, 9 * cm])
        routine_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        story.append(routine_table)
    else:
        story.append(Paragraph("No active routine yet.", styles["Normal"]))

    if adherence:
        story.append(Paragraph("Routine Adherence", styles["SectionHeading"]))
        story.append(
            _kv_table(
                [
                    ("Last 7 days", f"{adherence.get('7d')}%" if adherence.get("7d") is not None else "No data"),
                    ("Last 30 days", f"{adherence.get('30d')}%" if adherence.get("30d") is not None else "No data"),
                    ("Last 90 days", f"{adherence.get('90d')}%" if adherence.get("90d") is not None else "No data"),
                ]
            )
        )

    story.append(Paragraph("Recent Lifestyle Logs", styles["SectionHeading"]))
    if lifestyle_logs:
        log_rows = [["Date", "Sleep (h)", "Water (L)", "Stress"]] + [
            [
                log.logged_at.strftime("%d %b"),
                str(log.sleep_hours) if log.sleep_hours is not None else "—",
                str(log.water_intake_liters) if log.water_intake_liters is not None else "—",
                log.stress_level or "—",
            ]
            for log in lifestyle_logs
        ]
        log_table = Table(log_rows, colWidths=[4 * cm, 4 * cm, 4 * cm, 4 * cm])
        log_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        story.append(log_table)
    else:
        story.append(Paragraph("No lifestyle logs recorded yet.", styles["Normal"]))

    if recommendations:
        story.append(Paragraph("Product Recommendations", styles["SectionHeading"]))
        rec_rows = [["Product", "Brand", "Recommended by", "Note"]] + [
            [
                r.product.name if r.product else "—",
                r.product.brand if r.product else "—",
                "Your consultant",
                r.note or "—",
            ]
            for r in recommendations
        ]
        rec_table = Table(rec_rows, colWidths=[5 * cm, 3.5 * cm, 3.5 * cm, 4 * cm])
        rec_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        story.append(rec_table)

    doc.build(story)
    return buffer.getvalue()


def build_progress_report_pdf(db: Session, mongo_db: Database | None, user_id: uuid.UUID) -> bytes:
    """A focused progress report: score history, adherence, and the progress-photo timeline."""
    user = db.query(User).filter(User.id == user_id).first()
    history = assessment_service.get_assessment_history(db, user_id, limit=30)
    improvement = assessment_service.compute_improvement(db, user_id)
    adherence = progress_service.compute_adherence(db, mongo_db, user_id) if mongo_db is not None else {}
    photos = progress_service.list_progress_photos(db, user_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = _styles()
    story = []

    story.append(Paragraph("Progress Report", styles["ReportTitle"]))
    story.append(
        Paragraph(
            f"{user.full_name if user else 'Unknown user'} · Generated {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}",
            styles["ReportSubtitle"],
        )
    )

    story.append(Paragraph("Improvement Summary", styles["SectionHeading"]))
    if improvement:
        story.append(
            _kv_table(
                [
                    ("Starting score", f"{round(improvement['starting_score'])} / 100"),
                    ("Latest score", f"{round(improvement['latest_score'])} / 100"),
                    (
                        "Change",
                        f"{'+' if improvement['delta_points'] > 0 else ''}{improvement['delta_points']} pts "
                        f"({improvement['delta_percent']}%)",
                    ),
                    ("Trend", improvement["trend"]),
                    ("Tracking since", improvement["since"].strftime("%d %b %Y")),
                ]
            )
        )
    else:
        story.append(Paragraph("Not enough assessments yet to compute an improvement trend.", styles["Normal"]))

    story.append(Paragraph("Routine Adherence", styles["SectionHeading"]))
    story.append(
        _kv_table(
            [
                ("Last 7 days", f"{adherence.get('7d')}%" if adherence.get("7d") is not None else "No data"),
                ("Last 30 days", f"{adherence.get('30d')}%" if adherence.get("30d") is not None else "No data"),
                ("Last 90 days", f"{adherence.get('90d')}%" if adherence.get("90d") is not None else "No data"),
            ]
        )
    )

    story.append(Paragraph("Skin Health Score Timeline", styles["SectionHeading"]))
    if history:
        score_rows = [["Date", "Score"]] + [
            [a.created_at.strftime("%d %b %Y, %H:%M"), f"{round(a.overall_score)}"] for a in reversed(history)
        ]
        score_table = Table(score_rows, colWidths=[8 * cm, 4 * cm])
        score_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        story.append(score_table)
    else:
        story.append(Paragraph("No assessments recorded yet.", styles["Normal"]))

    story.append(Paragraph("Progress Photos", styles["SectionHeading"]))
    if photos:
        photo_rows = [["Date", "Tag", "Score at upload"]] + [
            [
                p.uploaded_at.strftime("%d %b %Y"),
                p.tag or "Untitled",
                f"{round(p.skin_health_score_at_upload)}" if p.skin_health_score_at_upload is not None else "—",
            ]
            for p in photos
        ]
        photo_table = Table(photo_rows, colWidths=[5 * cm, 5 * cm, 5 * cm])
        photo_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_COLOR),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        story.append(photo_table)
        story.append(
            Paragraph(
                "Photo image files are not embedded in this PDF — view them in-app under Progress.",
                styles["ReportSubtitle"],
            )
        )
    else:
        story.append(Paragraph("No progress photos uploaded yet.", styles["Normal"]))

    doc.build(story)
    return buffer.getvalue()


def build_history_excel(db: Session, user_id: uuid.UUID) -> bytes:
    """A multi-sheet Excel workbook: assessment history, lifestyle logs, and order history."""
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = "4F46E5"

    def _style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = __import__("openpyxl").styles.PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")

    ws1 = wb.active
    ws1.title = "Assessment History"
    ws1.append(["Date", "Overall Score", "Skin Type", "Primary Concern", "Condition", "Lifestyle", "Sleep", "Consistency", "Hydration"])
    for a in assessment_service.get_assessment_history(db, user_id, limit=200):
        ws1.append(
            [
                a.created_at.strftime("%Y-%m-%d %H:%M"),
                a.overall_score,
                a.skin_type,
                a.primary_concern,
                a.skin_condition_score,
                a.lifestyle_score,
                a.sleep_score,
                a.consistency_score,
                a.hydration_score,
            ]
        )
    _style_header(ws1, 9)

    ws2 = wb.create_sheet("Lifestyle Logs")
    ws2.append(["Date", "Sleep Hours", "Water (L)", "Exercise (min)", "Stress", "Smoking", "Alcohol", "Screen Time (h)"])
    for log in lifestyle_service.list_lifestyle_logs(db, user_id):
        ws2.append(
            [
                log.logged_at.strftime("%Y-%m-%d"),
                log.sleep_hours,
                log.water_intake_liters,
                log.exercise_minutes,
                log.stress_level,
                "Yes" if log.smoking else "No",
                "Yes" if log.alcohol else "No",
                log.screen_time_hours,
            ]
        )
    _style_header(ws2, 8)

    ws3 = wb.create_sheet("Order History")
    ws3.append(["Date", "Status", "Total (INR)", "Items"])
    for order in product_service.list_orders_for_user(db, user_id):
        item_summary = ", ".join(f"{item.product.name if item.product else '?'} x{item.quantity}" for item in order.items)
        ws3.append([order.created_at.strftime("%Y-%m-%d %H:%M"), order.status, float(order.total_amount), item_summary])
    _style_header(ws3, 4)

    for ws in (ws1, ws2, ws3):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
