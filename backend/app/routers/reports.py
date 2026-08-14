import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user

router = APIRouter(prefix="/api/reports", tags=["Reports & Export"])


def _gather_report_data(db: Session, user_id: int):
    profile = db.query(models.SkinProfile).filter(models.SkinProfile.user_id == user_id).first()
    assessment = (
        db.query(models.SkinAssessment)
        .filter(models.SkinAssessment.user_id == user_id)
        .order_by(models.SkinAssessment.created_at.desc())
        .first()
    )
    score = (
        db.query(models.SkinHealthScore)
        .filter(models.SkinHealthScore.user_id == user_id)
        .order_by(models.SkinHealthScore.computed_at.desc())
        .first()
    )
    routines = (
        db.query(models.Routine)
        .filter(models.Routine.user_id == user_id, models.Routine.is_active == True)  # noqa: E712
        .all()
    )
    return profile, assessment, score, routines


@router.get("/pdf")
def export_pdf_report(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    profile, assessment, score, routines = _gather_report_data(db, current_user.id)
    if not profile:
        raise HTTPException(status_code=400, detail="No profile data to report on yet.")

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 50

    def line(text, size=11, gap=18, bold=False):
        nonlocal y
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(50, y, text)
        y -= gap

    line(f"Skin Health Report - {current_user.name}", size=16, bold=True, gap=28)
    line(f"Skin Type: {profile.skin_type}   Age Group: {profile.age_group}")
    line(f"Concerns: {', '.join(profile.skin_concerns) or 'None recorded'}")
    line(f"Allergies: {', '.join(profile.allergies) or 'None recorded'}")
    y -= 10

    if assessment:
        line("Latest Skin Assessment", bold=True, size=13)
        line(f"Overall Condition Score: {assessment.overall_condition_score}/100")
        line(f"Prioritized Concerns: {', '.join(assessment.prioritized_concerns)}")
        for rf in assessment.risk_factors:
            line(f"- {rf}", size=10)
        y -= 10

    if score:
        line("Skin Health Score", bold=True, size=13)
        line(f"Overall: {score.overall_score}/100")
        line(f"Condition: {score.condition_score} | Lifestyle: {score.lifestyle_score} | "
             f"Sleep: {score.sleep_score} | Routine: {score.routine_consistency_score} | "
             f"Hydration: {score.hydration_score}", size=10)
        y -= 10

    if routines:
        line("Active Routines", bold=True, size=13)
        for r in routines:
            line(f"{r.routine_type.title()} routine ({len(r.steps)} steps)", size=10)

    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=skin_report_{current_user.id}.pdf"},
    )


@router.get("/excel")
def export_excel_report(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    from openpyxl import Workbook

    profile, assessment, score, routines = _gather_report_data(db, current_user.id)
    if not profile:
        raise HTTPException(status_code=400, detail="No profile data to report on yet.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Profile"
    ws.append(["Field", "Value"])
    ws.append(["Name", current_user.name])
    ws.append(["Skin Type", profile.skin_type])
    ws.append(["Age Group", profile.age_group])
    ws.append(["Concerns", ", ".join(profile.skin_concerns)])
    ws.append(["Allergies", ", ".join(profile.allergies)])
    ws.append(["Sleep Quality", profile.sleep_quality])
    ws.append(["Water Intake (L)", profile.water_intake_liters])

    if assessment:
        ws2 = wb.create_sheet("Assessment")
        ws2.append(["Concern", "Score"])
        for concern, val in assessment.condition_scores.items():
            ws2.append([concern, val])
        ws2.append(["Overall Condition Score", assessment.overall_condition_score])

    if score:
        ws3 = wb.create_sheet("Skin Health Score")
        ws3.append(["Component", "Score"])
        ws3.append(["Condition (35%)", score.condition_score])
        ws3.append(["Lifestyle (20%)", score.lifestyle_score])
        ws3.append(["Sleep (15%)", score.sleep_score])
        ws3.append(["Routine Consistency (20%)", score.routine_consistency_score])
        ws3.append(["Hydration (10%)", score.hydration_score])
        ws3.append(["Overall", score.overall_score])

    if routines:
        ws4 = wb.create_sheet("Routines")
        ws4.append(["Routine Type", "Order", "Category", "Product Category", "Instruction"])
        for r in routines:
            for step in r.steps:
                ws4.append([r.routine_type, step["order"], step["category"], step["product_category"], step["instruction"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=skin_report_{current_user.id}.xlsx"},
    )
