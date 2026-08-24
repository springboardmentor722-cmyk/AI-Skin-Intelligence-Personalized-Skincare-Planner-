import datetime
import io
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.storage import build_key, upload
from app.services.analytics.service import get_my_analytics
from app.services.recommendations.service import get_recommendations
from app.services.reports.models import ProgressReport, ReportSchedule
from app.services.reports.schemas import (
    ReportFormat,
    ReportScheduleCreate,
    ReportScheduleUpdate,
    ReportType,
)
from app.services.routines.service import get_or_generate_routines
from app.services.scores.service import get_latest_score
from app.services.skin_profile.models import SkinType
from app.services.skin_profile.service import get_current_profile

_STYLES = getSampleStyleSheet()


async def _profile_header_flowables(db: AsyncSession, user_id: str) -> list[Any]:
    profile = await get_current_profile(db, user_id)
    if profile is None:
        return []
    skin_type_name = "Unknown"
    result = await db.execute(
        select(SkinType.skin_type_name).where(SkinType.skin_type_id == profile.skin_type_id)
    )
    row = result.scalar_one_or_none()
    if row:
        skin_type_name = row
    generated_on = datetime.datetime.now(datetime.UTC).strftime("%B %d, %Y")
    return [
        Paragraph(f"Skin type: {skin_type_name}", _STYLES["Normal"]),
        Paragraph(f"Generated on: {generated_on}", _STYLES["Normal"]),
        Spacer(1, 12),
    ]


async def _assessment_flowables(db: AsyncSession, user_id: str) -> tuple[list[Any], str]:
    score = await get_latest_score(db, user_id)
    if score is None:
        return (
            [Paragraph("No skin assessment recorded yet.", _STYLES["Normal"])],
            "No assessment data available.",
        )
    rows = [
        ["Metric", "Score"],
        ["Skin Condition (35%)", f"{score.skin_condition_score or 0:.1f}"],
        ["Lifestyle (20%)", f"{score.lifestyle_score or 0:.1f}"],
        ["Routine Adherence (20%)", f"{score.routine_adherence_score or 0:.1f}"],
        ["Sleep Quality (15%)", f"{score.sleep_quality_score or 0:.1f}"],
        ["Hydration (10%)", f"{score.hydration_score or 0:.1f}"],
        ["Overall Score", f"{score.overall_score or 0:.1f}"],
    ]
    table = Table(rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
            ]
        )
    )
    summary = f"Overall skin health score: {score.overall_score or 0:.1f}/100."
    return ([Paragraph("Skin Assessment", _STYLES["Heading1"]), Spacer(1, 8), table], summary)


async def _progress_flowables(db: AsyncSession, user_id: str) -> tuple[list[Any], str]:
    analytics = await get_my_analytics(db, user_id, days=90)
    rows = [["Window", "Compliance %"]]
    rows.append(["7-day", f"{analytics.compliance.seven_day or 0:.0f}%"])
    rows.append(["30-day", f"{analytics.compliance.thirty_day or 0:.0f}%"])
    rows.append(["90-day", f"{analytics.compliance.ninety_day or 0:.0f}%"])
    table = Table(rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    summary = f"30-day routine compliance: {analytics.compliance.thirty_day or 0:.0f}%."
    return ([Paragraph("Progress Report", _STYLES["Heading1"]), Spacer(1, 8), table], summary)


async def _routine_flowables(db: AsyncSession, user_id: str) -> tuple[list[Any], str]:
    routines = await get_or_generate_routines(db, user_id)
    recs = await get_recommendations(db, user_id)
    flowables: list[Any] = [
        Paragraph("Routine & Recommendations", _STYLES["Heading1"]),
        Spacer(1, 8),
    ]
    for routine in routines:
        flowables.append(
            Paragraph(
                routine.routine_name or routine.routine_type or "Routine", _STYLES["Heading2"]
            )
        )
        for step in routine.steps:
            flowables.append(Paragraph(f"- {step.step_name or 'Step'}", _STYLES["Normal"]))
        flowables.append(Spacer(1, 8))
    if recs:
        flowables.append(Paragraph("Recommended Products", _STYLES["Heading2"]))
        for rec in recs[:5]:
            flowables.append(
                Paragraph(
                    f"- {rec.product.product_name or 'Product'} ({rec.match_percentage}% match)",
                    _STYLES["Normal"],
                )
            )
    summary = f"{len(routines)} routine(s), {len(recs)} recommendation(s)."
    return (flowables, summary)


_SECTION_BUILDERS = {
    "assessment": _assessment_flowables,
    "progress": _progress_flowables,
    "routine": _routine_flowables,
}


# --- Excel export (M4 audit fix, requirements PDF: "PDF export. Excel export.") ---
# Sibling row-builders, not a refactor of the flowables above: each calls the same
# underlying service function its PDF counterpart does and returns
# (header_row, data_rows, summary) instead of reportlab flowables. Small amount of
# duplication with the *_flowables functions above, same tradeoff this codebase
# already makes elsewhere rather than forcing one shared abstraction over two
# genuinely different output formats.


async def _assessment_rows(
    db: AsyncSession, user_id: str
) -> tuple[list[str], list[list[Any]], str]:
    score = await get_latest_score(db, user_id)
    header = ["Metric", "Score"]
    if score is None:
        return header, [], "No assessment data available."
    rows: list[list[Any]] = [
        ["Skin Condition (35%)", float(score.skin_condition_score or 0)],
        ["Lifestyle (20%)", float(score.lifestyle_score or 0)],
        ["Routine Adherence (20%)", float(score.routine_adherence_score or 0)],
        ["Sleep Quality (15%)", float(score.sleep_quality_score or 0)],
        ["Hydration (10%)", float(score.hydration_score or 0)],
        ["Overall Score", float(score.overall_score or 0)],
    ]
    summary = f"Overall skin health score: {score.overall_score or 0:.1f}/100."
    return header, rows, summary


async def _progress_rows(db: AsyncSession, user_id: str) -> tuple[list[str], list[list[Any]], str]:
    analytics = await get_my_analytics(db, user_id, days=90)
    header = ["Window", "Compliance %"]
    rows: list[list[Any]] = [
        ["7-day", float(analytics.compliance.seven_day or 0)],
        ["30-day", float(analytics.compliance.thirty_day or 0)],
        ["90-day", float(analytics.compliance.ninety_day or 0)],
    ]
    summary = f"30-day routine compliance: {analytics.compliance.thirty_day or 0:.0f}%."
    return header, rows, summary


async def _routine_rows(db: AsyncSession, user_id: str) -> tuple[list[str], list[list[Any]], str]:
    routines = await get_or_generate_routines(db, user_id)
    recs = await get_recommendations(db, user_id)
    header = ["Type", "Name", "Match %"]
    rows: list[list[Any]] = []
    for routine in routines:
        rows.append(["Routine", routine.routine_name or routine.routine_type or "Routine", None])
        for step in routine.steps:
            rows.append(["Step", step.step_name or "Step", None])
    for rec in recs[:5]:
        rows.append(["Product", rec.product.product_name or "Product", rec.match_percentage])
    summary = f"{len(routines)} routine(s), {len(recs)} recommendation(s)."
    return header, rows, summary


_ROW_BUILDERS = {
    "assessment": _assessment_rows,
    "progress": _progress_rows,
    "routine": _routine_rows,
}


def _build_xlsx(
    header: list[str],
    rows: list[list[Any]],
    *,
    sheet_title: str,
    profile_lines: list[tuple[str, Any]],
) -> bytes:
    """Header rows + a header row + data rows on one sheet. Empty `rows` still
    writes header-only (never errors on a brand-new user with no data yet)."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_title[:31]  # Excel's own sheet-name length limit.

    row_index = 1
    for label, value in profile_lines:
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        row_index += 1
    if profile_lines:
        row_index += 1  # blank separator row before the data table

    for column_index, title in enumerate(header, start=1):
        sheet.cell(row=row_index, column=column_index, value=title)
    row_index += 1
    for row in rows:
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
        row_index += 1

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def _profile_header_lines(db: AsyncSession, user_id: str) -> list[tuple[str, Any]]:
    profile = await get_current_profile(db, user_id)
    if profile is None:
        return []
    skin_type_name = "Unknown"
    result = await db.execute(
        select(SkinType.skin_type_name).where(SkinType.skin_type_id == profile.skin_type_id)
    )
    row = result.scalar_one_or_none()
    if row:
        skin_type_name = row
    return [
        ("Skin type", skin_type_name),
        ("Generated on", datetime.datetime.now(datetime.UTC).replace(tzinfo=None)),
    ]


async def _generate_xlsx_report(
    db: AsyncSession, user_id: str, report_type: ReportType, *, include_profile_header: bool
) -> ProgressReport:
    builder = _ROW_BUILDERS[report_type]
    header, rows, summary = await builder(db, user_id)
    profile_lines = await _profile_header_lines(db, user_id) if include_profile_header else []

    xlsx_bytes = _build_xlsx(
        header, rows, sheet_title=report_type.capitalize(), profile_lines=profile_lines
    )

    key = build_key(prefix="reports", owner_user_id=user_id, filename=f"{report_type}.xlsx")
    await upload(
        key,
        xlsx_bytes,
        allowed_content_types={"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )

    report = ProgressReport(
        user_id=user_id, report_type=report_type, format="xlsx", summary=summary, report_url=key
    )
    db.add(report)
    await db.commit()
    return report


async def generate_report(
    db: AsyncSession,
    user_id: str,
    report_type: ReportType,
    *,
    include_profile_header: bool,
    format: ReportFormat = "pdf",
) -> ProgressReport:
    if report_type not in _SECTION_BUILDERS:
        raise ValueError(f"Unknown report_type: {report_type!r}")

    if format == "xlsx":
        return await _generate_xlsx_report(
            db, user_id, report_type, include_profile_header=include_profile_header
        )

    builder = _SECTION_BUILDERS[report_type]
    flowables: list[Any] = []
    if include_profile_header:
        flowables.extend(await _profile_header_flowables(db, user_id))

    section_flowables, summary = await builder(db, user_id)
    flowables.extend(section_flowables)
    flowables.append(Spacer(1, 16))
    flowables.append(
        Paragraph(
            "This is general skincare guidance, not medical advice — check with a "
            "dermatologist for diagnosis or treatment.",
            _STYLES["Normal"],
        )
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    doc.build(flowables)
    pdf_bytes = buffer.getvalue()

    key = build_key(prefix="reports", owner_user_id=user_id, filename=f"{report_type}.pdf")
    await upload(key, pdf_bytes, allowed_content_types={"application/pdf"})

    report = ProgressReport(
        user_id=user_id, report_type=report_type, format="pdf", summary=summary, report_url=key
    )
    db.add(report)
    await db.commit()
    return report


async def list_my_schedules(db: AsyncSession, user_id: str) -> list[ReportSchedule]:
    result = await db.execute(select(ReportSchedule).where(ReportSchedule.user_id == user_id))
    return list(result.scalars().all())


async def create_schedule(
    db: AsyncSession, user_id: str, data: ReportScheduleCreate
) -> ReportSchedule:
    schedule = ReportSchedule(user_id=user_id, **data.model_dump())
    db.add(schedule)
    await db.commit()
    return schedule


async def _get_owned_schedule(db: AsyncSession, user_id: str, schedule_id: int) -> ReportSchedule:
    result = await db.execute(
        select(ReportSchedule).where(
            ReportSchedule.schedule_id == schedule_id, ReportSchedule.user_id == user_id
        )
    )
    schedule = result.scalar_one_or_none()
    if schedule is None:
        raise ValueError(f"Report schedule {schedule_id} not found for this user")
    return schedule


class ScheduleValidationError(Exception):
    """Raised when a PATCH would leave a schedule in a state the cron can never
    match (frequency='weekly' with no day_of_week, or 'monthly' with no
    day_of_month) — distinct from _get_owned_schedule's not-found ValueError so
    the router can map it to 422, not 404."""


async def update_schedule(
    db: AsyncSession, user_id: str, schedule_id: int, data: ReportScheduleUpdate
) -> ReportSchedule:
    schedule = await _get_owned_schedule(db, user_id, schedule_id)
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(schedule, field, value)
    if schedule.frequency == "weekly" and schedule.day_of_week is None:
        raise ScheduleValidationError("day_of_week is required when frequency is 'weekly'")
    if schedule.frequency == "monthly" and schedule.day_of_month is None:
        raise ScheduleValidationError("day_of_month is required when frequency is 'monthly'")
    await db.commit()
    return schedule


async def delete_schedule(db: AsyncSession, user_id: str, schedule_id: int) -> None:
    schedule = await _get_owned_schedule(db, user_id, schedule_id)
    await db.delete(schedule)
    await db.commit()
